"""
export_excel.py — Exportación Excel con formato para PediScan LCM
==================================================================
Genera dos hojas:
  1. Resultados individuales del evaluado actual (con colores por categoría)
  2. Tabla grupal acumulada de todos los evaluados

Requiere: openpyxl
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Paleta de colores por categoría
# ─────────────────────────────────────────────────────────────────────────────
FILL_PLANO         = PatternFill("solid", fgColor="CC3333")   # rojo
FILL_CAVO          = PatternFill("solid", fgColor="1A3E8C")   # azul oscuro
FILL_NORMAL        = PatternFill("solid", fgColor="1F7A1F")   # verde
FILL_INDETERMINADO = PatternFill("solid", fgColor="888888")   # gris
FILL_HEADER        = PatternFill("solid", fgColor="1C3557")   # azul marino
FILL_SUBHEADER     = PatternFill("solid", fgColor="2E6DA4")   # azul medio
FILL_ROW_ALT       = PatternFill("solid", fgColor="EEF4FB")   # azul muy claro

FONT_WHITE_BOLD  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
FONT_WHITE       = Font(color="FFFFFF", name="Calibri", size=10)
FONT_DARK        = Font(color="1C1C1C", name="Calibri", size=10)
FONT_DARK_BOLD   = Font(color="1C1C1C", bold=True, name="Calibri", size=10)
FONT_TITLE       = Font(color="1C3557", bold=True, name="Calibri", size=14)
FONT_SUBTITLE    = Font(color="444444", name="Calibri", size=10, italic=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
BORDER_MEDIUM_BOTTOM = Border(
    bottom=Side(style="medium", color="1C3557")
)


def _cat_fill(categoria):
    """Retorna el PatternFill correspondiente a la categoría."""
    cat = str(categoria).lower()
    if "plano"  in cat: return FILL_PLANO
    if "cavo"   in cat: return FILL_CAVO
    if "normal" in cat: return FILL_NORMAL
    return FILL_INDETERMINADO


def _style_cell(cell, fill=None, font=None, alignment=None, border=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _write_header_row(ws, row, cols, fill, font, height=22):
    """Escribe una fila de encabezado con estilo."""
    for col_idx, text in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        _style_cell(cell, fill=fill, font=font,
                    alignment=ALIGN_CENTER, border=BORDER_THIN)
    ws.row_dimensions[row].height = height


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1 — Resultados individuales
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_individual(wb, evaluado_id, fecha, feet_data, interpretation):
    ws = wb.create_sheet("Evaluado")

    # ── Título ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"] = "PediScan LCM — Análisis de huella plantar"
    _style_cell(ws["A1"], font=FONT_TITLE, alignment=ALIGN_LEFT)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Evaluado: {evaluado_id}     Fecha: {fecha}"
    _style_cell(ws["A2"], font=FONT_SUBTITLE, alignment=ALIGN_LEFT)
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 8  # separador

    # ── Encabezados de tabla ───────────────────────────────────────────────
    headers = [
        "Pie",
        "CSI (%)", "Cat. CSI",
        "SI",      "Cat. SI",
        "Clarke (°)", "Cat. Clarke",
        "AI",      "Cat. AI",
        "Categoría final", "Concordancia",
        "Nota AI*"
    ]
    _write_header_row(ws, 4, headers, FILL_HEADER, FONT_WHITE_BOLD, height=28)

    # ── Filas de datos ────────────────────────────────────────────────────
    pedigrafo_led = True   # siempre verdad para el LCM
    for r_idx, fd in enumerate(feet_data, start=5):
        met  = fd["metrics"]
        side = fd["foot"]["side"].capitalize()

        row_data = [
            side,
            met["CSI"],       met["cat_CSI"],
            met["SI"],        met["cat_SI"],
            met["Clarke"],    met["cat_Clarke"],
            met["AI"],        met["cat_AI"],
            met["cat_final"], f"{met['concordance']}/4",
            "Ver nota abajo" if pedigrafo_led else "—"
        ]

        fill_row = FILL_ROW_ALT if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            _style_cell(cell, fill=fill_row, font=FONT_DARK,
                        alignment=ALIGN_CENTER, border=BORDER_THIN)

        # Colorear celdas de categoría
        cat_cols = {3: met["cat_CSI"], 5: met["cat_SI"],
                    7: met["cat_Clarke"], 9: met["cat_AI"],
                    10: met["cat_final"]}
        for col, cat in cat_cols.items():
            cell = ws.cell(row=r_idx, column=col)
            _style_cell(cell, fill=_cat_fill(cat),
                        font=FONT_WHITE, alignment=ALIGN_CENTER, border=BORDER_THIN)

        ws.row_dimensions[r_idx].height = 20

    # ── Interpretación ────────────────────────────────────────────────────
    last_data_row = 4 + len(feet_data)
    sep_row = last_data_row + 2

    ws.merge_cells(f"A{sep_row}:L{sep_row}")
    ws[f"A{sep_row}"] = "Interpretación automática"
    _style_cell(ws[f"A{sep_row}"], fill=FILL_SUBHEADER,
                font=FONT_WHITE_BOLD, alignment=ALIGN_LEFT)
    ws.row_dimensions[sep_row].height = 20

    for i, line in enumerate(interpretation.strip().split("\n")):
        r = sep_row + 1 + i
        ws.merge_cells(f"A{r}:L{r}")
        ws[f"A{r}"] = line
        _style_cell(ws[f"A{r}"], font=FONT_DARK, alignment=ALIGN_LEFT)
        ws.row_dimensions[r].height = 16

    # ── Leyenda ───────────────────────────────────────────────────────────
    note_row = sep_row + len(interpretation.split("\n")) + 3
    ws.merge_cells(f"A{note_row}:L{note_row}")
    ws[f"A{note_row}"] = (
        "* Nota AI (pedígrafo LED): el Arch Index es elevado porque toda la planta "
        "contacta el vidrio. CSI y SI son los índices más confiables con este equipo."
    )
    _style_cell(ws[f"A{note_row}"],
                font=Font(color="885500", italic=True, size=9, name="Calibri"),
                alignment=ALIGN_LEFT)

    ws.merge_cells(f"A{note_row+1}:L{note_row+1}")
    ws[f"A{note_row+1}"] = (
        "Referencias: CSI: Forriol & Pascual (1990) Foot Ankle 11(2); "
        "SI: Staheli et al. (1987) JBJS Am 69(3); "
        "Clarke: Clarke (1933) Res Q AAHPE 4(3); "
        "AI: Cavanagh & Rodgers (1987) J Biomech 20(5)"
    )
    _style_cell(ws[f"A{note_row+1}"],
                font=Font(color="666666", italic=True, size=8, name="Calibri"),
                alignment=ALIGN_LEFT)

    # ── Leyenda de colores ────────────────────────────────────────────────
    legend_row = note_row + 3
    ws[f"A{legend_row}"] = "Leyenda:"
    _style_cell(ws[f"A{legend_row}"], font=FONT_DARK_BOLD)
    legend = [
        ("pie normal",   FILL_NORMAL,        "B"),
        ("pie plano",    FILL_PLANO,          "C"),
        ("pie cavo",     FILL_CAVO,           "D"),
        ("indeterminado",FILL_INDETERMINADO,  "E"),
    ]
    for label, fill, col in legend:
        cell = ws[f"{col}{legend_row}"]
        cell.value = label
        _style_cell(cell, fill=fill, font=FONT_WHITE,
                    alignment=ALIGN_CENTER, border=BORDER_THIN)

    ws.row_dimensions[legend_row].height = 18

    # ── Ancho de columnas ─────────────────────────────────────────────────
    col_widths = [14, 10, 16, 8, 14, 12, 16, 8, 14, 22, 14, 18]
    for i, cw in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = cw

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2 — Tabla grupal
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_grupal(wb, group_rows):
    ws = wb.create_sheet("Grupo")

    ws.merge_cells("A1:O1")
    ws["A1"] = "PediScan LCM — Tabla grupal de evaluación plantar"
    _style_cell(ws["A1"], font=FONT_TITLE, alignment=ALIGN_LEFT)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 8

    headers = [
        "ID", "Fecha",
        "CSI_D(%)", "Cat.CSI_D", "CSI_I(%)", "Cat.CSI_I",
        "SI_D",     "Cat.SI_D",  "SI_I",     "Cat.SI_I",
        "Clarke_D", "Cat.Cl_D",  "Clarke_I", "Cat.Cl_I",
        "Cat.Final_D", "Cat.Final_I",
    ]
    _write_header_row(ws, 3, headers, FILL_HEADER, FONT_WHITE_BOLD, height=30)

    cat_header_cols = [4, 6, 8, 10, 12, 14, 15, 16]

    for r_idx, row in enumerate(group_rows, start=4):
        fill_row = FILL_ROW_ALT if r_idx % 2 == 0 else None

        row_vals = [
            row.get("ID", ""),
            row.get("Fecha", ""),
            row.get("CSI_D", ""),
            row.get("Cat_D", ""),
            row.get("CSI_I", ""),
            row.get("Cat_I", ""),
            row.get("SI_D", ""),
            row.get("Cat_D", ""),
            row.get("SI_I", ""),
            row.get("Cat_I", ""),
            row.get("Clarke_D", ""),
            row.get("Cat_D", ""),
            row.get("Clarke_I", ""),
            row.get("Cat_I", ""),
            row.get("Cat_D", ""),
            row.get("Cat_I", ""),
        ]

        # Reconstruir desde las claves reales del group_row
        row_vals = [
            row.get("ID", ""),        row.get("Fecha", ""),
            row.get("CSI_D", ""),     row.get("Cat_D", ""),
            row.get("CSI_I", ""),     row.get("Cat_I", ""),
            row.get("SI_D", ""),      row.get("Cat_D", ""),
            row.get("SI_I", ""),      row.get("Cat_I", ""),
            row.get("Clarke_D", ""),  row.get("Cat_D", ""),
            row.get("Clarke_I", ""),  row.get("Cat_I", ""),
            row.get("Cat_D", ""),     row.get("Cat_I", ""),
        ]

        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            _style_cell(cell, fill=fill_row, font=FONT_DARK,
                        alignment=ALIGN_CENTER, border=BORDER_THIN)

        # Colorear columnas de categoría
        for col in cat_header_cols:
            cell = ws.cell(row=r_idx, column=col)
            if cell.value:
                _style_cell(cell, fill=_cat_fill(str(cell.value)),
                            font=FONT_WHITE, alignment=ALIGN_CENTER,
                            border=BORDER_THIN)

        ws.row_dimensions[r_idx].height = 18

    col_widths = [14, 12, 10, 16, 10, 16, 8, 14, 8, 14, 10, 16, 10, 16, 18, 18]
    for i, cw in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = cw

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(evaluado_id, fecha, feet_data,
                   interpretation, group_rows=None):
    """
    Genera un archivo Excel con dos hojas:
        - "Evaluado" — resultados individuales con colores
        - "Grupo"    — tabla grupal acumulada (si se proveen group_rows)

    Retorna:
        bytes del archivo .xlsx listos para st.download_button()
    """
    wb = Workbook()
    # Eliminar hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_individual(wb, evaluado_id, fecha, feet_data, interpretation)

    if group_rows:
        _sheet_grupal(wb, group_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
